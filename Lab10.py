#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Dec  7 20:01:56 2022

@author: lijiaqi
"""
import openpyxl

print("Hello, welcome to this program!")
print("The authors of this program are Jiaqi Li and Jiehao Xing.")

#Defining aggregate payment function
def total_payment_amount_function(passed_principal,n):
    i=0.1
    if n==1:
        total_payment_amount= "Paid in full"        
    else:
        numerater=float(passed_principal)*i*(1+i)**float(n)
        dinominator=(1+i)**float(n)-1
        total_payment_amount=(numerater/dinominator)*float(n)    
    return total_payment_amount

#Create initial lists
student_list=[]
tuition_code_list=[]
tuition_list=[]
sports_code_list=[]
sports_fee=[]
lunch_code_list=[]
lunch_fee=[]
status_code_list=[]
status_fee=[]
total_fees_and_tuition_list=[]
how_to_pay_list=[]
payment_number_list=[]
total_amount_list=[]

wb = openpyxl.load_workbook('student.xlsx') #opens excel file
sheet = wb['Sheet1'] #sets Sheet1 as active sheet

#Extract values and store in list
for r in range(1,int(sheet.max_row)+1):
    student_list.append(str(sheet.cell(row=r, column=1).value))
    tuition_code_list.append(str(sheet.cell(row=r, column=2).value))
    sports_code_list.append(str(sheet.cell(row=r, column=3).value))
    lunch_code_list.append(str(sheet.cell(row=r, column=4).value))
    status_code_list.append(str(sheet.cell(row=r, column=5).value))
    how_to_pay_list.append(str(sheet.cell(row=r, column=6).value))

print ("Do you want to input more student's information?")
keep_going=input("Input '1' to start; otherwise input '0'.")

#Adding students' information
while keep_going=='1': 

    #Asking the student's name
    student_name=input("What is the student's name?")
    student_list=student_list+[student_name]
        
    #Tuition
    print ("Is the student an in-district-student?")
    dist=input("Input '1' if the student is a in-district student otherwise please input '0'.")
    tuition_code_list=tuition_code_list+[dist]
        
    #Sports fee
    print ("Do the student plays sports?")
    sports=input("Input 'y' if the student plays a sport otherwise please input 'n'.")
    sports_code_list=sports_code_list+[sports]

    #Lunch fee
    print ("What type of lunch do the student eats?")
    lunch=input("Input '0' if the student brings lunch;'1' if the student eats standard lunch provided by school;'2' if the student eats premium lunch provided by school.")
    lunch_code_list=lunch_code_list+[lunch]


    #Student status
    print ("What is the student's status in school?")
    student_status=input("If the student is in initial year enter 'i'; If the student is in the guaduation year enter 'g'; Other please enter 'o'.")    
    status_code_list=status_code_list+[student_status]
     
        
    #Determine the way of payment and the amortized payment amount(round to two decimal points)
    print ("How will the student pay his or her bill?")
    payment_method=input("If the student wants to pay in whole right now, enter'w'; If the student wants to pay semi-annually, enter 's'; If the student wants to pay quaterly, enter 'q'; If the student pays monthly,enter 'm'." )
    how_to_pay_list=how_to_pay_list+[payment_method]
        
        
    print ("Do you want to keep going?")
    keep_going=input("Enter '1' to continue; enter '2' to end the calculation.")
  
#Transforming codes to values and store in new lists        
for i in range(len(tuition_code_list)):
    if tuition_code_list[i] =="1":
        tuition_list=tuition_list+[8000]
    elif tuition_code_list[i] =="0":
        tuition_list=tuition_list+[9500]
    
for i in range(len(sports_code_list)):
    if sports_code_list[i] =="n":
        sports_fee=sports_fee+[0]
    elif sports_code_list[i] =="y":
        sports_fee=sports_fee+[200]

for i in range(len(lunch_code_list)):
    if lunch_code_list[i]=="0":
        lunch_fee=lunch_fee+[0]
    elif lunch_code_list[i]=="1":
        lunch_fee=lunch_fee+[900]
    elif lunch_code_list[i]=="2":
        lunch_fee=lunch_fee+[1170]

for i in range(len(status_code_list)):
    if status_code_list[i]=="o":
        status_fee=status_fee+[0]
    elif status_code_list[i]=="i":
        status_fee=status_fee+[200]
    elif status_code_list[i]=="g":
        status_fee=status_fee+[100]

#Calculate the total fees and tuition to be collected from the student
for i in range(len(tuition_list)):
    total_fees_and_tuition_list.append(float(tuition_list[i])+float(sports_fee[i])+float(lunch_fee[i])+float(status_fee[i]))

for i in range(len(how_to_pay_list)):
    if how_to_pay_list[i]=="w":
        payment_number_list=payment_number_list+[1]
    elif how_to_pay_list[i]=="s":
        payment_number_list=payment_number_list+[2]
    elif how_to_pay_list[i]=="q":
        payment_number_list=payment_number_list+[4]
    elif how_to_pay_list[i]=="m":
        payment_number_list=payment_number_list+[12]

#Calculate the aggregate amount a student needs to pay after determining his or her payment method
for i in range(len(how_to_pay_list)):
    total_amount_list=total_amount_list+[total_payment_amount_function(total_fees_and_tuition_list[i],payment_number_list[i])]

#Update values in excel sheet
for r in range(len(total_amount_list)): 
    sheet.cell(row=r+1, column=1).value = student_list[r] 
    sheet.cell(row=r+1, column=2).value = tuition_code_list[r] 
    sheet.cell(row=r+1, column=3).value = sports_code_list[r] 
    sheet.cell(row=r+1, column=4).value = lunch_code_list[r] 
    sheet.cell(row=r+1, column=5).value = status_code_list[r]
    sheet.cell(row=r+1, column=6).value = how_to_pay_list[r]
    sheet.cell(row=r+1, column=7).value = total_amount_list[r]

wb.save('student.xlsx')



























    
    
